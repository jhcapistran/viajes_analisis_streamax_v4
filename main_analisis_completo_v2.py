"""
DETECCIÓN DE CAMBIOS DE CONDUCTOR
=======================================================

¿QUÉ HACE?
Detecta cambios de conductor durante un viaje usando dos métodos complementarios:
1. VISIÓN: Compara embeddings faciales entre frames consecutivos (distancia coseno)
2. FÍSICA: Valida si hubo tiempo físico suficiente para el cambio (basado en velocidad)

ESTRATEGIA DE DETECCIÓN:
- Comparación frame-a-frame estricta (visual_match = 0.5)
- visual_diff = 1.0: Todo lo que no sea match claro va al Turco (100% Recall)
- Gestión inteligente de IDs: Reutiliza IDs cuando el conductor regresa (patrón A→B→A)
- Prioridad por frames: IDs con más frames son más confiables (evita outliers)

INPUTS REQUERIDOS:
Excel con columnas:
   - embedding: Vector facial (JSON string o lista)
   - image_file: Nombre del archivo (formato: assetid_timestamp_extra.jpeg)
   - timestamp: Timestamp para visualización (no usado en cálculos)
   - speed: Velocidad del vehículo (km/h)
   - asset_id: ID del viaje
   - driver_id: ID del conductor (referencia/ground truth)
   - face_small, IPD_small: Flags de calidad para filtrado

OUTPUTS:
Excel con columnas adicionales:
   - PERSONA_ID: ID único asignado a cada conductor detectado
   - DECISION_SISTEMA: MISMO_CONDUCTOR | POSIBLE_CAMBIO | CAMBIO_CONFIRMADO
   - EXPLICACION: Justificación de la decisión
   - REQUIERE_VERIFICACION: SI/NO (para revisión del Turco)
   - Filas con cambios pintadas de verde

CONFIGURACIÓN OPTIMIZADA:
   Umbrales ajustados para balance óptimo entre recall y reducción de ruido:
   - visual_match = 0.5, visual_diff = 1.0
   - phys_impossible = 0.3, phys_possible = 0.85
   - Zona gris: alta_alta=0.73, media_alta=0.63, alta_media=0.78
   - t_maniobra dinámico: 30s (detenido), 60s (lento), 180s (movimiento)

RESULTADOS FINALES (101 cambios reales en dataset de 6,153 imágenes):
   - Recall: 100.00% (detecta los 101/101 cambios reales, sin falsos negativos)
   - Ahorro de Trabajo: 93.3% (Turco revisa 414 eventos de 6,153 originales)
   - Eficiencia de Alertas: 23.99% (~1 de cada 4 alertas es cambio real)
   - Chamba para el Turco: 414 eventos (94 cambios reales + 320 falsas alarmas)
   - Falsas Alarmas: 320 (precio por mantener 100% recall)
   - Filtrado de Calidad: 1,791 imágenes eliminadas automáticamente
   - Ratio IDs: 1.12 (57 IDs creados vs 51 reales)
   - Reutilización ID: 99.4% (asigna mismo ID cuando conductor regresa)
   - FP Confirmado: 2 (confirmaciones incorrectas de alta confianza)

"""

import os, json, math
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich import print as rprint

# --- CONFIGURACIÓN DE ARCHIVOS ---
INPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "trips_data_enriched.csv")

# --- CONFIGURACIÓN Y PARÁMETROS ---
# =============================================================================
# 🏆 CONFIGURACIÓN OPTIMIZADA - 02/01/2026
# =============================================================================
# Resultados actuales:
#   - 100% Recall (detecta todos los cambios reales: 101/101)
#   - Falsas Alarmas: 399 (-21% vs baseline)
#   - Chamba para el Turco: 494 (-24% vs baseline)
#   - Eficiencia de Alertas: 20.2%
#   - Ratio IDs: 1.12 (57 IDs vs 51 reales)
#   - Reutilización: 99.4% (asigna correctamente mismo ID cuando conductor regresa)
#   - Falsos Positivos Confirmados: 2
#
# ESTRATEGIA:
# - Comparación estricta frame-a-frame (visual_match = 0.5)
# - visual_diff = 1.0: Todo no-match claro va a revisión
# - Zona gris (0.5-1.0): Lógica combinada visual + física:
#   * P_física > 0.85 y dist > 0.6 → Alerta
#   * P_física > 0.3 y dist > 0.75 → Alerta
#   * Resto → Mismo conductor
# - t_maniobra dinámico según velocidad:
#   * v < 3.6 km/h: 30 seg (detenido)
#   * v < 18 km/h: 60 seg (lento)
#   * v ≥ 18 km/h: 180 seg (normal)
# - Gestión inteligente de IDs:
#   * Reutiliza IDs con umbral relajado (0.7) cuando conductor regresa
#   * Prioriza IDs con más frames (más confiables, evita outliers)
#
# PARÁMETROS CLAVE:
#   visual_match = 0.5   (match claro)
#   visual_diff = 1.0    (diferencia clara)
#   phys_impossible = 0.3  (teletransporte)
#   phys_possible = 0.85   (tiempo suficiente)
#   id_reuse_match_threshold = 0.7  (reutilización IDs)
#
# =============================================================================

CONSTANTS = {
    # --- FÍSICA DEL VEHÍCULO ---
    'a_decel': 2.0,       # Desaceleración (m/s²) - para calcular tiempo de frenado
    'a_accel': 0.5,       # Aceleración (m/s²) - para calcular tiempo de arranque
    't_maniobra': 90,     # Tiempo mínimo para cambio de conductor (segundos) - Antes 180
    't_maniobra_stationary': 15, # Tiempo de maniobra con vehiculo practicamente detenido
    't_maniobra_slow': 45,       # Tiempo de maniobra con velocidad muy baja
    'k': 0.1,             # Suavizado de curva logística (transición gradual)
    
    # --- COMPARACIÓN VISUAL (FRAME A FRAME) ---
    'visual_match': 0.5,   # Umbral para considerar misma persona (comparación estricta)
    'visual_diff': 1.0,    # Umbral confirmación (1.0 = todo no-match va al Turco)
    'gray_margin': 0.15,   # Margen de zona gris (entre match y diff)
    'gray_zone_high': 0.73,         # Distancia alta en zona gris con fisica favorable
    'gray_zone_medium': 0.63,       # Distancia media en zona gris con fisica favorable
    'gray_zone_medium_high': 0.78,  # Distancia alta en zona gris con fisica dudosa
    
    # --- GESTIÓN DE IDs ---
    'id_reuse_match_threshold': 0.7, # Umbral relajado para reutilizar IDs (P95 intra-conductor)
    
    # --- VALIDACIÓN FÍSICA ---
    'phys_impossible': 0.1, # Umbral para imposibilidad física (< 0.1 = muy probable teletransporte) - Antes 0.3
    'phys_possible': 0.85,  # Umbral para posibilidad alta (> 0.85 = muy probable que hubo tiempo)
    
    # --- FILTROS DE CALIDAD (BASURA IN -> BASURA OUT) ---
    'filter_face_small': True, # Si la cara es minúscula, la ignoramos.
    'filter_ipd_small': True,  # Si la resolución entre ojos es mala, bye.
    'min_images_per_asset': 10, # Mínimo de imágenes para considerar un viaje válido.
    'min_stationary_speed': 7.0, # Velocidad (km/h) por debajo de la cual se considera detenido para filtrado
}

# --- CONTADOR GLOBAL DE IDs ---
GLOBAL_ID_COUNTER = 0

def get_new_id():
    """Genera un nuevo ID consecutivo de 6 dígitos."""
    global GLOBAL_ID_COUNTER
    GLOBAL_ID_COUNTER += 1
    return f"{GLOBAL_ID_COUNTER:06d}"

def load_emb(emb_str):
    """Intenta sacar el vector numérico de la celda del Excel o el Parquet."""
    if emb_str is None:
        return None
        
    try:
        # Si ya es un array de numpy o algo parecido, no hacemos el check de truthiness directo
        if isinstance(emb_str, (np.ndarray, list)):
            return np.array(emb_str, dtype=float)

        # Si es pandas scalar NA
        if pd.isna(emb_str):
            return None
            
        # Si es texto, hay que parsear el JSON
        if isinstance(emb_str, str):
            data = json.loads(emb_str)
            return np.array(data, dtype=float)
            
        return None
            
        return None
    except Exception as e:
        # Si falla, pues ni modo, retornamos None
        return None
def process_asset_group(df_asset, asset_name="Unknown"):
    """
    Procesa un viaje/asset completo y detecta cambios de conductor.
    
    Estrategia:
    1. Filtrar frames de baja calidad
    2. Comparar embeddings frame-a-frame (distancia coseno)
    3. Validar cambios con análisis físico (velocidad/tiempo)
    4. Gestionar IDs: reutilizar cuando conductor regresa, crear nuevo cuando cambia
    """
    print(f"   -> Procesando viaje/asset: {asset_name} ({len(df_asset)} filas)")
    
    # 1. Filtrado de frames de baja calidad
    df_clean = df_asset.copy()
    
    # Filtro 1: Descartar caras demasiado pequeñas
    if CONSTANTS['filter_face_small'] and 'face_small' in df_clean.columns:
        df_clean = df_clean[df_clean['face_small'] == False]
        
    # Filtro 2: Descartar frames con mala resolución IPD (distancia entre pupilas)
    if CONSTANTS['filter_ipd_small'] and 'IPD_small' in df_clean.columns:
        df_clean = df_clean[df_clean['IPD_small'] == 0.0]

    # --- EXTRACCIÓN DE TIMESTAMP ---
    img_col = 'image_file' if 'image_file' in df_clean.columns else None
    if not img_col:
        print("      [!] Error: No se encontró columna de nombre de archivo (image_file).")
        return []

    def get_ts_from_file(val):
        try:
            parts = str(val).split('_')
            if len(parts) >= 2:
                return int(parts[1]) // 1000
            return np.nan
        except: return np.nan

    df_clean['ts_seconds'] = df_clean[img_col].apply(get_ts_from_file)
    df_clean = df_clean.dropna(subset=['ts_seconds'])
    df_clean['ts_seconds'] = df_clean['ts_seconds'].astype(int)
    df_clean = df_clean.sort_values('ts_seconds').reset_index(drop=True)

    # --- SPEED FILTERING (Top-down) ---
    speeds = df_clean['speed'].values if 'speed' in df_clean.columns else np.zeros(len(df_clean))
    min_stat = CONSTANTS['min_stationary_speed']
    keep_mask = speeds >= min_stat
    df_clean = df_clean[keep_mask].reset_index(drop=True)

    # --- SKIPPING INITIAL FRAME ---
    # Saltamos el primer frame de cada viaje para evitar falsas alarmas de cold start
    if len(df_clean) > 1:
        df_clean = df_clean.iloc[1:].reset_index(drop=True)
    
    if df_clean.empty or len(df_clean) <= CONSTANTS['min_images_per_asset']:
        print(f"      [!] Asset {asset_name} sin suficientes datos válidos ({len(df_clean)} frames).")
        return []

    # --- INICIALIZACIÓN DE METADATA (CON DATOS FILTRADOS) ---
    records = []
    a_decel = CONSTANTS['a_decel']
    a_accel = CONSTANTS['a_accel']
    t_maniobra = CONSTANTS['t_maniobra']
    k = CONSTANTS['k']
    last_valid_idx = 0

    # DEBUG: Mostrar ejemplo de conversión de timestamp del primer frame REAL
    ejemplo = df_clean.iloc[0]
    print(f"      🔍 [DEBUG TIMESTAMP] Primer frame analizado: {ejemplo[img_col]} -> Segundos: {ejemplo['ts_seconds']} (Speed: {ejemplo['speed']})")

    # Inicializamos el ID con una nueva persona para el primer frame real
    current_person_id = get_new_id()
    id_embeddings_pool = {current_person_id: []}
    
    first_row = df_clean.iloc[0]
    emb_first = load_emb(first_row.get('embedding'))
    if emb_first is not None:
        id_embeddings_pool[current_person_id].append(emb_first)
        
    records = [{
        'ASSET_ID': asset_name,
        'PERSONA_ID': current_person_id,
        'DRIVER_ID': first_row.get('driver_id', 'N/A'),
        'IDENTITY_ID': first_row.get('identity_id', ''),
        'GS_PATH': first_row.get('gs_path', ''),
        'EMPTY_CABIN': first_row.get('empty_cabin', ''),
        'ARCHIVO': first_row.get(img_col, 'N/A'),
        'SOURCE_IMAGE': first_row.get('source_image', 'N/A'),
        'TIMESTAMP': first_row.get('timestamp', 'N/A'),
        'VELOCIDAD': float(first_row.get('speed', 0.0)),
        'DISTANCIA_VISUAL': None,
        'PROBABILIDAD_FISICA': 1.0,
        'DECISION_SISTEMA': 'INICIO_VIAJE',
        'EXPLICACION': 'Primer frame del viaje (post-filtrado)',
        'REQUIERE_VERIFICACION': 'NO',
        'DELTA_T_SEC': 0
    }]
    
    id_centroids = {} 
    if id_embeddings_pool[current_person_id]:
        id_centroids[current_person_id] = np.mean(id_embeddings_pool[current_person_id], axis=0)
    
    # A darle vueltas al asunto
    # El filtrado y el skip ya se hicieron arriba.
    for i in range(1, len(df_clean)):
        prev = df_clean.iloc[last_valid_idx]
        cur = df_clean.iloc[i]

        # Velocidades (de km/h a m/s porque física usa SI)
        speed_prev = float(prev.get('speed', 0.0))
        speed_cur = float(cur.get('speed', 0.0))
        v_prev = speed_prev / 3.6
        v_curr = speed_cur / 3.6

        # NOTA: Ya no saltamos 0→0 porque el filtrado previo eliminó los ceros intermedios.
        # Los ceros que quedan son frontera y deben procesarse.

        last_valid_idx = i

        # Tiempos (Ya están en segundos)
        prev_ts = prev['ts_seconds']
        cur_ts = cur['ts_seconds']
        delta_t = cur_ts - prev_ts

        # Embeddings
        emb_cur = load_emb(cur.get('embedding'))
        emb_prev = load_emb(prev.get('embedding'))

        # --- CÁLCULO VISUAL FRAME A FRAME ---
        dist_prev = None
        if emb_prev is not None and emb_cur is not None:
            denom = (np.linalg.norm(emb_prev) * np.linalg.norm(emb_cur))
            if denom > 0:
                dist_prev = 1.0 - float(np.dot(emb_prev, emb_cur) / denom)
        main_distance = dist_prev

        # --- GESTIÓN INTELIGENTE DE IDs (REUTILIZACIÓN) ---
        # El sistema busca reutilizar IDs existentes cuando:
        # 1. Es falso positivo: no hubo cambio real, mantener mismo ID
        # 2. Conductor regresa: patrón A→B→A (cambio temporal), reutilizar ID de A
        # 3. Si no hay match: crear nuevo ID
        #
        # ESTRATEGIA:
        # - Compara embedding actual con centroides de todos los IDs previos
        # - Usa umbral relajado (0.7) vs frame-a-frame (0.5) para reutilización
        # - Prioriza IDs con más frames (más confiables, evita outliers de 1-2 frames)
        reused_id = None
        best_reuse_dist = float('inf')
        best_reuse_frames = 0
        if emb_cur is not None and len(id_centroids) > 0:
            id_reuse_match_threshold = CONSTANTS['id_reuse_match_threshold']
            
            for pid, centroid in id_centroids.items():
                # Calcular distancia del frame actual al centroide del ID
                pid_embeddings = id_embeddings_pool.get(pid, [])
                denom = (np.linalg.norm(centroid) * np.linalg.norm(emb_cur))
                if denom > 0:
                    dist_to_centroid = 1.0 - float(np.dot(centroid, emb_cur) / denom)
                    if dist_to_centroid < id_reuse_match_threshold:
                        num_frames = len(pid_embeddings)
                        # Priorizar: 1) Más frames (más confiable), 2) Menor distancia
                        if num_frames > best_reuse_frames or (num_frames == best_reuse_frames and dist_to_centroid < best_reuse_dist):
                            best_reuse_dist = dist_to_centroid
                            best_reuse_frames = num_frames
                            reused_id = pid

        # --- VALIDACIÓN FÍSICA (TIEMPO DISPONIBLE PARA CAMBIO) ---
        # Calcula si hubo tiempo físico suficiente para cambio de conductor
        # basado en velocidades y tiempos de frenado/arranque
        t_frenado = v_prev / a_decel         # Tiempo para frenar completamente
        t_arranque = v_curr / a_accel        # Tiempo para volver a velocidad actual
        
        # t_maniobra dinámico: Si el vehículo ya está detenido o casi detenido,
        # el cambio es más rápido (30-60 seg). Si va a alta velocidad, necesita más tiempo (180 seg).
        # Usamos la velocidad máxima entre ambos frames como referencia
        v_max = max(v_prev, v_curr)
        if v_max < 1.0:  # < 3.6 km/h (prácticamente detenido)
            t_maniobra_dinamico = CONSTANTS['t_maniobra_stationary']
        elif v_max < 5.0:  # < 18 km/h (velocidad muy baja, casi detenido)
            t_maniobra_dinamico = CONSTANTS['t_maniobra_slow']
        else:  # Velocidad normal/alta
            t_maniobra_dinamico = t_maniobra
        
        t_req = t_frenado + t_maniobra_dinamico + t_arranque  # Tiempo total requerido
        t_sobra = delta_t - t_req            # Tiempo sobrante (puede ser negativo)
        
        # Sigmoide para probabilidad física
        P_fisica = 1.0 / (1.0 + math.exp(-k * t_sobra))

        # --- EL VEREDICTO ---
        decision = 'MISMO_CONDUCTOR'
        requiere_ver = 'NO'
        explanation = ''
        
        visual_match = CONSTANTS['visual_match']
        gray_margin = CONSTANTS['gray_margin']
        visual_diff = CONSTANTS['visual_diff']
        # anchor_limit eliminado
        if main_distance is None:
            decision = 'INDETERMINADO'
            explanation = 'Sin embedding no hay paraíso (no se pudo calcular)'
            requiere_ver = 'SI'
        else:
            # CHEQUEO DE SEGURIDAD: Drift eliminado. Solo frame a frame.
            is_drift = False

            # CHEQUEO DE SEGURIDAD: ¿Cambio brusco frame a frame?
            is_sudden = False
            if dist_prev is not None and dist_prev > visual_diff:
                is_sudden = True
                explanation += f' | Salto brusco vs anterior ({dist_prev:.3f})'

            # --- ZONA 1: IDENTIDAD (Match Claro) ---
            # Embeddings muy similares (dist < 0.5)
            if main_distance < visual_match:
                # Validar contra anomalías (drift gradual o saltos bruscos de velocidad)
                if is_drift or is_sudden:
                     decision = 'POSIBLE_CAMBIO'
                     requiere_ver = 'SI'
                     explanation = f'Centroide OK ({main_distance:.3f}) pero: ' + explanation
                else:
                    decision = 'MISMO_CONDUCTOR'
                    explanation = f'Se parecen un buen ({main_distance:.3f})'
            
            # --- ZONA 2: DIFERENCIA CLARA (No Match) ---
            # Embeddings muy diferentes (dist > 1.0)
            elif main_distance > visual_diff:
                # Validar con física: ¿Hubo tiempo para cambiar?
                if P_fisica > CONSTANTS['phys_impossible']:
                    decision = 'CAMBIO_CONFIRMADO'
                    requiere_ver = 'NO'  # Alta confianza
                    explanation = f'Cambio claro visual ({main_distance:.3f}) y físico'
                else:
                    decision = 'MISMO_CONDUCTOR'
                    explanation = f'Caras distintas pero físicamente imposible (teletransporte)'
            
            # --- ZONA 3: ZONA GRIS (Incertidumbre) ---
            # Embeddings en zona intermedia (0.5 <= dist <= 1.0)
            else:
                # Estrategia combinada: Visual * Física
                # Cuanto más diferente visualmente + más tiempo físico = más sospechoso
                
                if P_fisica > CONSTANTS['phys_possible']:
                    # Física muy posible (>0.85) + zona media visual
                    # Entre más alta la distancia visual, más sospechoso
                    if main_distance > CONSTANTS['gray_zone_high']:
                        # Zona media-alta visual + física posible = muy sospechoso
                        decision = 'POSIBLE_CAMBIO'
                        requiere_ver = 'SI'
                        explanation = f'Zona media-alta ({main_distance:.3f}), física posible (P={P_fisica:.2f})'
                    elif main_distance > CONSTANTS['gray_zone_medium']:
                        # Zona media visual + física posible = sospechoso
                        decision = 'POSIBLE_CAMBIO'
                        requiere_ver = 'SI'
                        explanation = f'Zona media ({main_distance:.3f}), física posible (P={P_fisica:.2f})'
                    else:
                        # Zona baja de incertidumbre + física posible = probablemente mismo conductor
                        decision = 'MISMO_CONDUCTOR'
                        explanation = f'Zona media-baja ({main_distance:.3f}), física posible pero visual similar (P={P_fisica:.2f})'
                
                elif P_fisica > CONSTANTS['phys_impossible']:
                    # Física dudosa (0.3-0.85)
                    if main_distance > CONSTANTS['gray_zone_medium_high']:
                        # Visual muy diferente + física dudosa = revisar
                        decision = 'POSIBLE_CAMBIO'
                        requiere_ver = 'SI'
                        explanation = f'Zona alta ({main_distance:.3f}), física dudosa (P={P_fisica:.2f})'
                    else:
                        # Visual no tan diferente + física dudosa = mismo conductor
                        decision = 'MISMO_CONDUCTOR'
                        explanation = f'Zona media ({main_distance:.3f}), física dudosa (P={P_fisica:.2f})'
                
                else:
                    # P_fisica < 0.3: Físicamente muy improbable (casi imposible)
                    decision = 'MISMO_CONDUCTOR'
                    explanation = f'Zona media ({main_distance:.3f}), física imposible (P={P_fisica:.2f})'

        # --- ASIGNACIÓN DE PERSONA_ID ---
        # Estrategia:
        # - CAMBIO_CONFIRMADO/POSIBLE_CAMBIO: Intentar reutilizar ID existente, si no crear nuevo
        # - MISMO_CONDUCTOR: Mantener ID actual, o crear uno si es el primer frame
        if decision in ['CAMBIO_CONFIRMADO', 'POSIBLE_CAMBIO']:
            if reused_id is not None:
                current_person_id = reused_id  # Reutilizar ID (conductor regresó)
            else:
                current_person_id = get_new_id()  # Crear nuevo ID (conductor nuevo)
        elif decision == 'MISMO_CONDUCTOR':
            if current_person_id is None:
                current_person_id = get_new_id()  # Primer frame del viaje

        # --- ACTUALIZAR MEMORIA DE IDs ---
        # Mantener historial de embeddings y centroides para cada ID
        if emb_cur is not None:
            if current_person_id not in id_embeddings_pool:
                id_embeddings_pool[current_person_id] = []
            id_embeddings_pool[current_person_id].append(emb_cur)
            # Recalcular centroide como promedio de todos los embeddings del ID
            arr = np.stack(id_embeddings_pool[current_person_id])
            id_centroids[current_person_id] = np.mean(arr, axis=0)

        # --- REGISTRAR RESULTADO ---
        records.append({
            'ASSET_ID': asset_name,
            'PERSONA_ID': current_person_id,
            'DRIVER_ID': cur.get('driver_id', 'N/A'),
            'IDENTITY_ID': cur.get('identity_id', ''),
            'GS_PATH': cur.get('gs_path', ''),
            'EMPTY_CABIN': cur.get('empty_cabin', ''),
            'ARCHIVO': cur.get(img_col, 'N/A'),
            'SOURCE_IMAGE': cur.get('source_image', 'N/A'),
            'TIMESTAMP': cur.get('timestamp', 'N/A'),
            'VELOCIDAD': speed_cur,
            'DISTANCIA_VISUAL': round(main_distance, 4) if main_distance is not None else None,
            'PROBABILIDAD_FISICA': round(P_fisica, 4),
            'DECISION_SISTEMA': decision,
            'EXPLICACION': explanation,
            'REQUIERE_VERIFICACION': requiere_ver,
            'DELTA_T_SEC': round(delta_t, 2)
        })
        
    return records

def load_and_process(input_file=None):
    """Carga el CSV de entrada y corre el pipeline de detección por viaje.

    Devuelve el DataFrame de salida (out_df) con una fila por frame procesado,
    o None si no se pudo cargar o no salió ningún registro. Reutilizable por
    main() y por scripts auxiliares (ej. reporte HTML de misses).
    """
    input_file = input_file or INPUT_FILE

    # 1. Cargar el CSV
    try:
        df = pd.read_csv(input_file)
        print(f"✅ CSV cargado: {len(df)} filas.")
    except Exception as e:
        print(f"❌ Error cargando CSV: {e}")
        return None

    # Validar que traiga embeddings
    if 'embedding' not in df.columns:
        print("❌ ERROR CRÍTICO: No hay columna 'embedding'.")
        print("   Sin embeddings no hay magia.")
        return None

    # Quedarnos solo con filas que traen embedding
    df = df[df['embedding'].notna() & (df['embedding'].astype(str).str.strip() != '')]
    print(f"✅ Filas con embedding válido: {len(df)}.")

    # Generar 'image_file' esperado por process_asset_group: asset_id + "_" + nombre del archivo
    df['image_file'] = df['asset_id'].astype(str) + "_" + df['gs_path'].apply(lambda p: os.path.basename(str(p)))

    all_records = []

    # 2. Procesar (Por viaje si hay ID, si no, todo junto)
    if 'asset_id' in df.columns:
        assets = df['asset_id'].unique()
        print(f"ℹ️ Se encontraron {len(assets)} viajes distintos.")
        for asset in assets:
            df_asset = df[df['asset_id'] == asset]
            results = process_asset_group(df_asset, str(asset))
            all_records.extend(results)
    else:
        print("ℹ️ No vi columna 'asset_id', así que asumo que es todo del mismo viaje.")
        results = process_asset_group(df, "Single_Asset")
        all_records.extend(results)

    if not all_records:
        print("⚠️ No salió nada. ¿Filtros muy agresivos o datos vacíos?")
        return None

    return pd.DataFrame(all_records)


def compute_gt_changes(out_df):
    """Calcula los cambios reales (GT) usando IDENTITY_ID, exigiendo que la
    nueva identidad se sostenga >= 2 frames (para ignorar parpadeos de 1 frame).

    Devuelve una lista de dicts, uno por cada cambio GT encontrado:
      asset_id, prev_identity, new_identity, cur_start (posición dentro del
      grupo del asset, 0-indexed), prev_run_end (última posición de la
      identidad previa), detected (bool).
    """
    changes = []
    for asset_id, group in out_df.groupby('ASSET_ID', sort=False):
        identity = group['IDENTITY_ID'].astype(str).str.strip().tolist()
        decision = group['DECISION_SISTEMA'].tolist()
        n = len(identity)

        # Run-length encode la secuencia de identidades: (valor, inicio, fin_inclusive)
        runs = []
        i = 0
        while i < n:
            j = i
            while j + 1 < n and identity[j + 1] == identity[i]:
                j += 1
            runs.append((identity[i], i, j))
            i = j + 1

        # Solo cuentan como "estado real" las identidades válidas sostenidas >= 2 frames
        valid_runs = [r for r in runs if r[0] not in ('', 'nan') and (r[2] - r[1] + 1) >= 2]

        for k in range(1, len(valid_runs)):
            prev_val, _, prev_end = valid_runs[k - 1]
            cur_val, cur_start, _ = valid_runs[k]
            if prev_val != cur_val:
                detected = decision[cur_start] in ('POSIBLE_CAMBIO', 'CAMBIO_CONFIRMADO')
                changes.append({
                    'asset_id': asset_id,
                    'prev_identity': prev_val,
                    'new_identity': cur_val,
                    'cur_start': cur_start,
                    'prev_run_end': prev_end,
                    'detected': detected,
                })
    return changes


def main():
    # Generamos el nombre de salida automáticamente
    base, ext = os.path.splitext(INPUT_FILE)
    OUTPUT_FILE = f"{base}_resultados.xlsx"

    print(f"🚀 INICIANDO ANÁLISIS")
    print(f"📂 Entrada: {INPUT_FILE}")
    print(f"💾 Salida:  {OUTPUT_FILE}")
    print("-" * 40)

    out_df = load_and_process(INPUT_FILE)
    if out_df is None:
        return

    # --- AGREGAR ALERTAS DE IDs SOSPECHOSOS ---
    # Columnas para flagging de IDs problemáticos
    out_df['ID_FLAG'] = ''
    out_df['RECOMENDACION'] = ''
    
    # Detectar mini-IDs (≤2 frames)
    persona_frame_counts = out_df.groupby(['ASSET_ID', 'PERSONA_ID']).size()
    mini_ids = persona_frame_counts[persona_frame_counts <= 2]
    
    for (asset_id, persona_id), count in mini_ids.items():
        mask = (out_df['ASSET_ID'] == asset_id) & (out_df['PERSONA_ID'] == persona_id)
        alerta = "🔴 MINI-ID (1 frame)" if count == 1 else "🟡 MINI-ID (2 frames)"
        out_df.loc[mask, 'ID_FLAG'] = alerta
        out_df.loc[mask, 'RECOMENDACION'] = f"Probable falso positivo - Revisar consolidación"
    
    # Detectar conductores fragmentados (1 conductor real → múltiples PERSONA_IDs)
    # Solo si tenemos DRIVER_ID disponible
    # IMPORTANTE: Solo flagging en las transiciones (donde CAMBIA el ID), no todo el conductor
    if 'DRIVER_ID' in out_df.columns:
        for asset_id in out_df['ASSET_ID'].unique():
            # IMPORTANTE: NO usar reset_index(drop=True) para mantener los índices correctos de out_df
            asset_df = out_df[out_df['ASSET_ID'] == asset_id]
            
            # Mapeo: driver_id real → lista de PERSONA_IDs
            driver_to_personas = {}
            for driver_id, group in asset_df.groupby('DRIVER_ID'):
                if str(driver_id) not in ['N/A', 'nan', '']:
                    personas = group['PERSONA_ID'].unique()
                    if len(personas) > 1:
                        driver_to_personas[driver_id] = sorted(personas)
            
            # Flag de fragmentación SOLO en puntos de transición
            if driver_to_personas:
                asset_indices = asset_df.index.tolist()
                
                for i in range(1, len(asset_df)):
                    curr_row_idx = asset_indices[i]
                    
                    prev_row = asset_df.iloc[i-1]
                    curr_row = asset_df.iloc[i]
                    
                    prev_driver = prev_row.get('DRIVER_ID', 'N/A')
                    curr_driver = curr_row.get('DRIVER_ID', 'N/A')
                    prev_persona = prev_row.get('PERSONA_ID', '')
                    curr_persona = curr_row.get('PERSONA_ID', '')
                    
                    # Si es el MISMO conductor pero DIFERENTE ID → transición de fragmentación
                    if (prev_driver == curr_driver and 
                        prev_driver in driver_to_personas and 
                        prev_persona != curr_persona):
                        
                        # Flag solo en el frame actual (inicio de nuevo ID)
                        personas_list = driver_to_personas[prev_driver]
                        personas_str = ", ".join(str(p) for p in personas_list)
                        
                        out_df.loc[curr_row_idx, 'ID_FLAG'] = f"🔵 CAMBIO ID (mismo conductor)"
                        out_df.loc[curr_row_idx, 'RECOMENDACION'] = f"Consolidar con IDs: {personas_str}"
    
    out_df.to_excel(OUTPUT_FILE, index=False)
    print(f"\n✅ Resultados guardados en: {OUTPUT_FILE}")
    print(f"   Incluye columnas: ID_FLAG, RECOMENDACION (para alertas de IDs sospechosos)")

    # 4. Ponerle colorcitos (Verde para alertas)
    try:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        
        # Buscar dónde quedó la decisión
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            decision_col_idx = header.index('DECISION_SISTEMA') + 1
        except ValueError:
            decision_col_idx = None

        if decision_col_idx:
            fill_alert = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # Verde
            
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell_decision = row[decision_col_idx - 1]
                val = cell_decision.value
                if val in ['POSIBLE_CAMBIO', 'CAMBIO_CONFIRMADO']:
                    for cell in row:
                        cell.fill = fill_alert
            
            wb.save(OUTPUT_FILE)
            print("🎨 Colores aplicados. Se ve bonito.")
    except Exception as e:
        print(f"⚠️ Falló la pintada de colores: {e}")

    # Resumen Final con Rich
    console = Console()
    
    # --- RESUMEN GLOBAL ---
    n_total = len(out_df)
    n_alertas = len(out_df[out_df['DECISION_SISTEMA'].isin(['POSIBLE_CAMBIO', 'CAMBIO_CONFIRMADO'])])
    n_turco = len(out_df[out_df['REQUIERE_VERIFICACION'] == 'SI'])
    
    n_total_ids = out_df.groupby(['ASSET_ID', 'PERSONA_ID']).ngroups
    
    table_global = Table(title="📊 RESUMEN GLOBAL", show_header=True, header_style="bold magenta")
    table_global.add_column("Métrica", style="cyan", no_wrap=True)
    table_global.add_column("Cantidad", justify="right", style="green")
    table_global.add_column("Comentario", style="italic")

    table_global.add_row("Total Eventos Analizados", str(n_total), "Todo lo que pasó por el filtro")
    table_global.add_row("Total Personas (IDs)", str(n_total_ids), "Conductores distintos detectados")
    table_global.add_row("Alertas Detectadas", str(n_alertas), "Cambios o posibles cambios")
    table_global.add_row("Chamba para el Turco", str(n_turco), "[bold red]Lo que hay que revisar a mano[/bold red]")

    console.print("\n")
    console.print(table_global)

    # --- MÉTRICAS DE DETECCIÓN (GT = identity_id del CSV) ---
    # GT: hay cambio real cuando la nueva identidad se sostiene al menos 2 frames seguidos
    # (evita contar parpadeos de 1 frame como cambio real).
    gt_changes = compute_gt_changes(out_df)
    gt_total = len(gt_changes)
    gt_detected = sum(1 for c in gt_changes if c['detected'])

    tp_positions_by_asset = {}
    for c in gt_changes:
        if c['detected']:
            tp_positions_by_asset.setdefault(c['asset_id'], set()).add(c['cur_start'])

    fp_count = 0
    fp_confirmado = 0
    fp_posible = 0
    for asset_id, group in out_df.groupby('ASSET_ID', sort=False):
        decision = group['DECISION_SISTEMA'].tolist()
        tp_positions = tp_positions_by_asset.get(asset_id, set())
        for pos, dec in enumerate(decision):
            if dec in ('POSIBLE_CAMBIO', 'CAMBIO_CONFIRMADO') and pos not in tp_positions:
                fp_count += 1
                if dec == 'CAMBIO_CONFIRMADO':
                    fp_confirmado += 1
                else:
                    fp_posible += 1

    recall_pct = (gt_detected / gt_total * 100) if gt_total else 0.0
    precision_pct = (gt_detected / (gt_detected + fp_count) * 100) if (gt_detected + fp_count) else 0.0
    # Precisión "estricta": solo penaliza CAMBIO_CONFIRMADO (compromiso firme del
    # sistema). Los POSIBLE_CAMBIO son solo advertencias para que el Turco revise
    # y muchas veces terminan resolviéndose como MISMO_CONDUCTOR, así que no
    # deberían pesar igual que un cambio confirmado incorrecto.
    precision_estricta_pct = (gt_detected / (gt_detected + fp_confirmado) * 100) if (gt_detected + fp_confirmado) else 0.0

    table_metrics = Table(title="🎯 MÉTRICAS DE DETECCIÓN (vs identity_id, sostenida >= 2 frames)", show_header=True, header_style="bold cyan")
    table_metrics.add_column("Métrica", style="cyan", no_wrap=True)
    table_metrics.add_column("Valor", justify="right", style="green")
    table_metrics.add_column("Comentario", style="italic")

    table_metrics.add_row("Cambios Reales (GT)", str(gt_total), "Cambios de identity_id sostenidos >= 2 frames")
    table_metrics.add_row("Cambios Detectados", str(gt_detected), "GT con alerta del sistema en la misma fila")
    table_metrics.add_row("Recall", f"{recall_pct:.2f}%", "% de cambios reales detectados")
    table_metrics.add_row("Falsos Positivos (total)", str(fp_count), "Alertas que no coinciden con un cambio real")
    table_metrics.add_row("  ↳ Falsos POSIBLE_CAMBIO", str(fp_posible), "Solo advertencia, no compromiso firme")
    table_metrics.add_row("  ↳ Falsos CAMBIO_CONFIRMADO", str(fp_confirmado), "Compromiso firme incorrecto (más grave)")
    table_metrics.add_row("Precisión", f"{precision_pct:.2f}%", "% de alertas (todas) que fueron cambio real")
    table_metrics.add_row("Precisión Estricta", f"{precision_estricta_pct:.2f}%", "Solo penaliza CAMBIO_CONFIRMADO incorrectos")

    console.print("\n")
    console.print(table_metrics)

    # --- RESUMEN POR ASSET (VIAJE) ---
    if 'ASSET_ID' in out_df.columns:
        table_assets = Table(title="🚚 RESUMEN POR VIAJE (ASSET)", show_header=True, header_style="bold blue")
        table_assets.add_column("Asset ID", style="yellow")
        table_assets.add_column("Eventos", justify="right")
        table_assets.add_column("IDs Únicos", justify="right", style="magenta")
        table_assets.add_column("Alertas", justify="right", style="red")
        table_assets.add_column("Revisión Manual", justify="right", style="bold red")
        
        assets = out_df['ASSET_ID'].unique()
        for asset in assets:
            df_a = out_df[out_df['ASSET_ID'] == asset]
            n_a_total = len(df_a)
            n_a_alertas = len(df_a[df_a['DECISION_SISTEMA'].isin(['POSIBLE_CAMBIO', 'CAMBIO_CONFIRMADO'])])
            n_a_turco = len(df_a[df_a['REQUIERE_VERIFICACION'] == 'SI'])
            n_a_ids = df_a['PERSONA_ID'].nunique()
            
            table_assets.add_row(str(asset), str(n_a_total), str(n_a_ids), str(n_a_alertas), str(n_a_turco))
        
        console.print("\n")
        console.print(table_assets)
    
    # --- DETECCIÓN DE IDs SOSPECHOSOS (FRAGMENTACIÓN POTENCIAL) ---
    # Análisis post-procesamiento: encontrar IDs que probablemente pertenezcan al mismo conductor
    console.print("\n[bold cyan]🔍 ANÁLISIS DE IDs SOSPECHOSOS (Fragmentación Potencial)[/bold cyan]\n")
    
    # --- ANÁLISIS SIMPLE: Contar frames por PERSONA_ID (igual que evaluacion_desempeno.py) ---
    console.print("\n[bold magenta]📌 ANÁLISIS DE IDs MINI (Posibles Falsos Positivos)[/bold magenta]\n")
    
    # 1. Detectar IDs mini (≤2 frames)
    mini_ids_by_asset = {}
    for asset_id in out_df['ASSET_ID'].unique():
        asset_df = out_df[out_df['ASSET_ID'] == asset_id]
        persona_frame_counts = asset_df['PERSONA_ID'].value_counts()
        
        mini_ids = {pid: count for pid, count in persona_frame_counts.items() if count <= 2}
        if mini_ids:
            mini_ids_by_asset[asset_id] = mini_ids
    
    if mini_ids_by_asset:
        total_mini = sum(len(mini_ids) for mini_ids in mini_ids_by_asset.values())
        tabla_mini_ids = Table(title="⚠️ IDs con Muy Pocos Frames (Posibles Falsos Positivos)", 
                               show_header=True, header_style="bold magenta")
        tabla_mini_ids.add_column("Asset", style="cyan")
        tabla_mini_ids.add_column("PERSONA_ID", style="yellow", justify="right")
        tabla_mini_ids.add_column("# Frames", justify="right", style="red")
        tabla_mini_ids.add_column("Alerta", style="bold red")
        tabla_mini_ids.add_column("Sugerencia para Turco", style="italic")
        
        count_shown = 0
        for asset_id in sorted(mini_ids_by_asset.keys()):
            for pid, frame_count in sorted(mini_ids_by_asset[asset_id].items()):
                alerta = "🔴 1 frame" if frame_count == 1 else "🟡 2 frames"
                sugerencia = "Muy probablemente falso positivo - Revisar" if frame_count == 1 else "Probablemente ruido - Revisar"
                tabla_mini_ids.add_row(
                    str(asset_id),
                    str(pid),
                    str(frame_count),
                    alerta,
                    sugerencia
                )
                count_shown += 1
                if count_shown >= 30:
                    break
            if count_shown >= 30:
                break
        
        console.print(tabla_mini_ids)
        console.print(f"\n[bold yellow]⚠️ Encontrados {total_mini} IDs con ≤2 frames (probable ruido/falsos positivos)[/bold yellow]")
        console.print("[italic]💡 Estos IDs deberían revisarse: pueden ser cambios reales, pero muchos son ruido.[/italic]\n")
    else:
        console.print("[green]✅ No hay IDs con ≤2 frames. Buena señal.[/green]\n")

    # Mensaje final
    if n_turco > 0:
        console.print(Panel(f"[bold yellow]¡OJO![/bold yellow] El Turco tiene [bold red]{n_turco}[/bold red] eventos para revisar en total. ¡A darle!", title="⚠️ ALERTA DE CHAMBA", border_style="red"))
    else:
        console.print(Panel("[bold green]¡Todo limpio![/bold green] El Turco se puede ir temprano hoy.", title="✅ SIN NOVEDAD", border_style="green"))

if __name__ == "__main__":
    main()
